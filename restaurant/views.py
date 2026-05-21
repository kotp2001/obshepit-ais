from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import datetime, timedelta
import json
import sqlite3
import os
from .models import Category, Dish, Table, Order, OrderItem, Booking, MaintenanceLog, Employee
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.contrib import messages
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ===== АВТОРИЗАЦИЯ =====
def custom_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = CustomLoginForm()
    
    return render(request, 'registration/login.html', {'form': form})

def custom_logout(request):
    logout(request)
    return redirect('login')

# ===== ГЛАВНАЯ И DASHBOARD =====
def home(request):
    return render(request, 'home.html')

@login_required
def dashboard(request):
    # Определяем роль пользователя
    try:
        employee = request.user.employee
        role = employee.role
    except Employee.DoesNotExist:
        if request.user.is_superuser:
            role = 'admin'
        else:
            role = 'waiter'
    
    if role == 'waiter':
        return redirect('waiter_hall')
    elif role == 'cook':
        return redirect('kitchen_view')
    elif role == 'admin' or request.user.is_superuser:
        return redirect('admin_panel')
    
    return render(request, 'dashboard.html', {'role': role})

# ===== ЗАЛ ОФИЦИАНТА =====
@login_required
def waiter_hall(request):
    tables = Table.objects.all().order_by('number')
    return render(request, 'waiter/hall.html', {'tables': tables})

@login_required
def create_order(request, table_id):
    table = get_object_or_404(Table, id=table_id)
    categories = Category.objects.all().order_by('order')
    
    if request.method == 'POST':
        data = json.loads(request.body)
        items = data.get('items', [])
        
        if not items:
            return JsonResponse({'error': 'Заказ пуст'}, status=400)
        
        order = Order.objects.create(
            table=table,
            status='active',
            total_amount=0
        )
        
        total = 0
        for item in items:
            dish = get_object_or_404(Dish, id=item['dish_id'])
            quantity = item['quantity']
            price = dish.price * quantity
            total += price
            
            OrderItem.objects.create(
                order=order,
                dish=dish,
                quantity=quantity,
                price=price,
                status='pending'
            )
        
        order.total_amount = total
        order.save()
        
        # Меняем статус стола
        table.status = 'occupied'
        table.save()
        
        return JsonResponse({'success': True, 'order_id': order.id})
    
    return render(request, 'waiter/create_order.html', {
        'table': table,
        'categories': categories
    })

@login_required
def pay_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        payment_method = request.POST.get('payment_method')
        order.payment_method = payment_method
        order.status = 'paid'
        order.save()
        
        # Освобождаем стол
        table = order.table
        table.status = 'free'
        table.save()
        
        return redirect('order_receipt', order_id=order.id)
    
    return render(request, 'waiter/pay_order.html', {'order': order})

@login_required
def order_receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'waiter/receipt.html', {'order': order})

# ===== КУХНЯ (ПОВАР) =====
@login_required
def kitchen_view(request):
    orders = Order.objects.filter(status='active').prefetch_related('order_items__dish').order_by('created_at')
    return render(request, 'cook/kitchen.html', {'orders': orders})

@login_required
def update_order_item_status(request, order_item_id):
    if request.method == 'POST':
        order_item = get_object_or_404(OrderItem, id=order_item_id)
        status = request.POST.get('status')
        
        if status in ['pending', 'cooking', 'ready']:
            order_item.status = status
            order_item.save()
            
            return JsonResponse({'success': True, 'status': status})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def mark_order_ready(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        order.status = 'ready'
        order.save()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

# ===== АДМИН ПАНЕЛЬ =====
@login_required
@staff_member_required
def admin_panel(request):
    # Статистика
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    active_orders = Order.objects.filter(status='active').count()
    total_tables = Table.objects.count()
    free_tables = Table.objects.filter(status='free').count()
    
    # График по дням
    orders_by_day = Order.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id'),
        revenue=Sum('total_amount')
    ).order_by('date')
    
    # Популярные блюда
    popular_dishes = OrderItem.objects.values('dish__name').annotate(
        total_quantity=Sum('quantity')
    ).order_by('-total_quantity')[:5]
    
    context = {
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'active_orders': active_orders,
        'total_tables': total_tables,
        'free_tables': free_tables,
        'orders_by_day': list(orders_by_day),
        'popular_dishes': list(popular_dishes),
    }
    
    return render(request, 'admin_panel.html', context)

# ===== API =====
def api_tables(request):
    tables = Table.objects.all().values('id', 'number', 'status', 'seats', 'position_x', 'position_y')
    return JsonResponse({'tables': list(tables)})

def api_menu(request):
    categories = Category.objects.all().prefetch_related('dish_set').order_by('order')
    data = []
    for category in categories:
        cat_data = {
            'id': category.id,
            'name': category.name,
            'icon': category.icon,
            'dishes': []
        }
        for dish in category.dish_set.filter(is_available=True):
            cat_data['dishes'].append({
                'id': dish.id,
                'name': dish.name,
                'price': float(dish.price),
                'description': dish.description or '',
            })
        data.append(cat_data)
    return JsonResponse({'categories': data})

def api_create_order(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    data = json.loads(request.body)
    table_id = data.get('table_id')
    items = data.get('items', [])
    
    if not table_id or not items:
        return JsonResponse({'error': 'Missing data'}, status=400)
    
    table = get_object_or_404(Table, id=table_id)
    
    order = Order.objects.create(
        table=table,
        status='active',
        total_amount=0
    )
    
    total = 0
    for item in items:
        dish = get_object_or_404(Dish, id=item['dish_id'])
        quantity = item['quantity']
        price = dish.price * quantity
        total += price
        
        OrderItem.objects.create(
            order=order,
            dish=dish,
            quantity=quantity,
            price=price,
            status='pending'
        )
    
    order.total_amount = total
    order.save()
    
    table.status = 'occupied'
    table.save()
    
    return JsonResponse({'success': True, 'order_id': order.id})

def api_pay_order(request, order_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    order = get_object_or_404(Order, id=order_id)
    data = json.loads(request.body)
    
    order.payment_method = data.get('payment_method', 'cash')
    order.status = 'paid'
    order.save()
    
    table = order.table
    table.status = 'free'
    table.save()
    
    return JsonResponse({'success': True})

def api_update_order_item_status(request, item_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    order_item = get_object_or_404(OrderItem, id=item_id)
    data = json.loads(request.body)
    
    status = data.get('status')
    if status in ['pending', 'cooking', 'ready']:
        order_item.status = status
        order_item.save()
        return JsonResponse({'success': True, 'status': status})
    
    return JsonResponse({'error': 'Invalid status'}, status=400)

# ===== СТРАНИЦЫ ДОКУМЕНТАЦИИ =====
def help_manual(request):
    return render(request, 'help_manual.html')

def regulations(request):
    return render(request, 'regulations.html')

@login_required
@staff_member_required
def maintenance_log_view(request):
    logs = MaintenanceLog.objects.all().order_by('-date')
    return render(request, 'maintenance_log.html', {'logs': logs})

@login_required
@staff_member_required
def backup_view(request):
    """Страница резервного копирования"""
    return render(request, 'backup.html')

# ===== ЭКСПОРТ В EXCEL =====
@login_required
@staff_member_required
def export_orders_excel(request):
    orders = Order.objects.filter(status='paid').select_related('table').order_by('-created_at')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Заголовки
    headers = ['ID', 'Дата', 'Стол', 'Сумма', 'Способ оплаты', 'Статус']
    ws.append(headers)
    
    # Стиль заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    for order in orders:
        ws.append([
            order.id,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            f"Стол {order.table.number}",
            order.total_amount,
            order.get_payment_method_display() if order.payment_method else '-',
            order.get_status_display()
        ])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=orders_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@staff_member_required
def export_dishes_excel(request):
    dishes = Dish.objects.select_related('category').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Блюда"
    
    headers = ['Категория', 'Название', 'Цена', 'Описание', 'Доступно']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for dish in dishes:
        ws.append([
            dish.category.name if dish.category else '-',
            dish.name,
            dish.price,
            dish.description or '-',
            'Да' if dish.is_available else 'Нет'
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=dishes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response
